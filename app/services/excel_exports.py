from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.models import Ticket, Visitor, VisitorAnswer
from app.services.tickets import get_project_detailed_stats

FUNNEL_FIELD_LABELS = {
    "started_bot": "Нажали Start в боте",
    "started_registration": "Начали регистрацию (отправили имя)",
    "left_contact": "Ввели номер телефона",
    "registration_completed": "Завершили регистрацию",
    "tickets_issued": "Выдано билетов",
    "opened_my_ticket": "Открыли раздел «Мой билет»",
    "tickets_annulled": "Аннулировано билетов",
    "attended_qr_scan": "Пришли на мероприятие (QR отсканирован)",
    "lottery_participants": "Участники розыгрыша",
}


def _fit_columns(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def _format_datetime(value) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def build_lottery_tickets_excel(db: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лотерейные билеты"
    sheet.append(["Лотерейный код"])

    lottery_codes = db.scalars(
        select(Ticket.lottery_code)
        .where(
            Ticket.lottery_code.is_not(None),
            Ticket.lottery_code != "",
        )
        .order_by(Ticket.created_at.asc())
    ).all()
    for lottery_code in lottery_codes:
        sheet.append([lottery_code])

    _fit_columns(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _funnel_status_for_visitor(visitor: Visitor) -> str:
    """Определяет статус пользователя в воронке по данным Visitor и связанного Ticket."""
    if visitor.ticket is None:
        if visitor.is_registration_completed:
            return "Билет аннулирован"
        return "Регистрация не завершена"
    t = visitor.ticket
    if t.is_activated and t.lottery_code:
        return "Пришёл на мероприятие (QR), участник розыгрыша"
    if t.is_activated:
        return "Пришёл на мероприятие (QR)"
    return "Билет выдан"


def build_analytics_excel(db: Session) -> bytes:
    workbook = Workbook()

    funnel_sheet = workbook.active
    funnel_sheet.title = "Воронка"
    stats = get_project_detailed_stats(db=db)
    funnel_sheet.append(["Показатель", "Значение"])
    for key, label in FUNNEL_FIELD_LABELS.items():
        funnel_sheet.append([label, stats.get(key, 0)])
    _fit_columns(funnel_sheet)

    users_sheet = workbook.create_sheet("Пользователи по статусу")
    users_sheet.append(
        [
            "Телеграм ID",
            "Имя пользователя",
            "Полное имя",
            "Статус в воронке",
            "Номер билета",
            "Лотерейный код",
            "Билет активирован",
            "Дата активации билета",
            "Дата создания",
            "Дата обновления",
        ]
    )

    visitors_with_tickets = db.scalars(
        select(Visitor)
        .options(joinedload(Visitor.ticket))
        .order_by(Visitor.created_at.asc())
    ).all()

    for visitor in visitors_with_tickets:
        ticket = visitor.ticket
        status = _funnel_status_for_visitor(visitor)
        users_sheet.append(
            [
                visitor.telegram_id,
                visitor.username or "",
                visitor.full_name or "",
                status,
                ticket.ticket_number if ticket else "",
                ticket.lottery_code if ticket else "",
                "Да" if ticket and ticket.is_activated else "Нет",
                _format_datetime(ticket.activated_at) if ticket else "",
                _format_datetime(visitor.created_at),
                _format_datetime(visitor.updated_at),
            ]
        )
    _fit_columns(users_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
